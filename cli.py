#!/usr/bin/env python3

#!/usr/bin/env python3
"""
Command-Line Interface for Earnings Call Analyzer
Phase 2A Enhanced - Includes deception analysis commands
"""
import click
import json
import logging
from pathlib import Path
from config.settings import settings
from config.logging_config import setup_logging

# Initialize logging
setup_logging()


@click.group()
@click.version_option(version="2.0.0-phase2a")
def cli():
	"""Earnings Call Analyzer - Comprehensive financial communication analysis"""
	pass
	
	
@cli.command()
@click.argument('transcript_file', type=click.Path(exists=True))
@click.option('--output', '-o', help='Output JSON file path', default=None)
@click.option('--no-llm', is_flag=True, help='Disable LLM features')
@click.option('--with-deception', is_flag=True, default=True, help='Include Phase 2A deception analysis (default: enabled)')
@click.option('--summary', '-s', is_flag=True, help='Print summary to console')
def analyze(transcript_file, output, no_llm, with_deception, summary):
	"""
	Analyze an earnings call transcript (Phase 1 + Phase 2A)
	
	Example:
		earnings-analyzer analyze transcript.txt --with-deception --summary
	"""
	from src.analysis.aggregator import EarningsCallAnalyzer
	
	# Initialize analyzer
	analyzer = EarningsCallAnalyzer(
		use_llm_features=not no_llm,
		enable_deception_analysis=with_deception
	)
	
	# Analyze
	results = analyzer.analyze_transcript(transcript_file)
	
	# Save results
	if output:
		analyzer.save_results(results, output)
	else:
		# Default output path
		input_path = Path(transcript_file)
		output_path = input_path.with_suffix('.results.json')
		analyzer.save_results(results, str(output_path))
		
	# Print summary
	if summary:
		analyzer.print_summary(results)
		
		
@cli.command()
@click.argument('results_file', type=click.Path(exists=True))
def summary(results_file):
	"""
	Print summary from a results JSON file
	
	Example:
		earnings-analyzer summary results.json
	"""
	from src.analysis.aggregator import EarningsCallAnalyzer, ComprehensiveAnalysisResult
	from dataclasses import fields
	
	# Load results
	with open(results_file, 'r') as f:
		data = json.load(f)
		
	# Reconstruct result object (simplified - in production use proper deserialization)
	# For now, just print key metrics
	click.echo(f"\n{'='*80}")
	click.echo(f"EARNINGS CALL ANALYSIS SUMMARY")
	click.echo(f"Company: {data.get('company_name')} | Quarter: {data.get('quarter')} {data.get('year')}")
	click.echo(f"{'='*80}")
	
	# Phase 1 metrics
	click.echo("\n📊 PHASE 1: CORE METRICS")
	click.echo("-" * 80)
	if 'overall_sentiment' in data:
		sent = data['overall_sentiment']
		click.echo(f"Sentiment: {sent.get('hybrid_label')} ({sent.get('hybrid_sentiment_score'):.2f})")
	if 'overall_complexity' in data:
		comp = data['overall_complexity']
		click.echo(f"Complexity: {comp.get('complexity_level')} ({comp.get('composite_score'):.0f}/100)")
	if 'overall_numerical' in data:
		num = data['overall_numerical']
		click.echo(f"Numerical Transparency: {num.get('numeric_transparency_score'):.2f}%")
		
	# Phase 2A metrics
	if data.get('deception_risk'):
		click.echo("\n⚠️  PHASE 2A: DECEPTION RISK ASSESSMENT")
		click.echo("-" * 80)
		risk = data['deception_risk']
		click.echo(f"Overall Risk: {risk.get('risk_level')} ({risk.get('overall_risk_score'):.0f}/100)")
		click.echo(f"Confidence: {risk.get('confidence'):.0%}")
		
		if risk.get('triggered_flags'):
			click.echo(f"\nTriggered Flags:")
			for flag in risk['triggered_flags'][:5]:
				click.echo(f"  • {flag}")
				
	# Red flags
	if data.get('red_flags'):
		click.echo("\n🚩 RED FLAGS")
		click.echo("-" * 80)
		for flag in data['red_flags']:
			click.echo(f"  • {flag}")
			
	# Strengths
	if data.get('strengths'):
		click.echo("\n✅ STRENGTHS")
		click.echo("-" * 80)
		for strength in data['strengths']:
			click.echo(f"  • {strength}")
			
	click.echo("\n" + "="*80)
	
	
@cli.command()
@click.argument('transcript_file', type=click.Path(exists=True))
@click.option('--output', '-o', help='Output file path', default=None)
def deception(transcript_file, output):
	"""
	Run ONLY deception analysis on a transcript (Phase 2A)
	
	Example:
		earnings-analyzer deception transcript.txt -o deception_report.json
	"""
	from src.analysis.aggregator import EarningsCallAnalyzer
	
	click.echo("\n🔍 Running deception-focused analysis...")
	
	# Initialize with deception only
	analyzer = EarningsCallAnalyzer(
		use_llm_features=True,
		enable_deception_analysis=True
	)
	
	# Analyze
	results = analyzer.analyze_transcript(transcript_file)
	
	# Extract deception-specific results
	deception_report = {
		'company': results.company_name,
		'quarter': results.quarter,
		'year': results.year,
		'timestamp': results.timestamp,
		'deception_risk': results.deception_risk.__dict__ if results.deception_risk else None,
		'evasiveness': results.evasiveness_scores.__dict__ if results.evasiveness_scores else None,
		'qa_analysis': [qa.__dict__ for qa in results.qa_analysis] if results.qa_analysis else []
	}
	
	# Save
	if output:
		output_path = Path(output)
	else:
		input_path = Path(transcript_file)
		output_path = input_path.with_suffix('.deception.json')
		
	with open(output_path, 'w') as f:
		json.dump(deception_report, f, indent=2, default=str)
		
	click.echo(f"\n✓ Deception analysis saved to: {output_path}")
	
	# Print quick summary
	if results.deception_risk:
		click.echo(f"\nRisk Level: {results.deception_risk.risk_level}")
		click.echo(f"Risk Score: {results.deception_risk.overall_risk_score:.0f}/100")
		click.echo(f"Flags: {len(results.deception_risk.triggered_flags)}")
		
		
@cli.command()
@click.argument('results_file', type=click.Path(exists=True))
@click.option('--metric', '-m', help='Specific metric to display', default=None)
def inspect(results_file, metric):
	"""
	Inspect detailed metrics from analysis results
	
	Example:
		earnings-analyzer inspect results.json --metric deception_risk
	"""
	with open(results_file, 'r') as f:
		data = json.load(f)
		
	if metric:
		if metric in data:
			click.echo(f"\n{metric.upper()}:")
			click.echo(json.dumps(data[metric], indent=2))
		else:
			click.echo(f"Metric '{metric}' not found in results")
			click.echo(f"\nAvailable metrics:")
			for key in data.keys():
				click.echo(f"  - {key}")
	else:
		click.echo("\nAll metrics:")
		for key in data.keys():
			click.echo(f"  - {key}")
			
			
@cli.command()
@click.argument('directory', type=click.Path(exists=True))
@click.option('--format', '-f', type=click.Choice(['json', 'csv']), default='json')
@click.option('--with-deception', is_flag=True, default=True)
def batch(directory, format, with_deception):
	"""
	Batch process all transcripts in a directory
	
	Example:
		earnings-analyzer batch ./transcripts/ --with-deception
	"""
	from src.analysis.aggregator import EarningsCallAnalyzer
	import glob
	
	dir_path = Path(directory)
	transcript_files = list(dir_path.glob('*.txt')) + list(dir_path.glob('*.md'))
	
	if not transcript_files:
		click.echo(f"No transcript files found in {directory}")
		return
	
	click.echo(f"\n📦 Batch processing {len(transcript_files)} transcripts...")
	
	analyzer = EarningsCallAnalyzer(
		use_llm_features=True,
		enable_deception_analysis=with_deception
	)
	
	results_list = []
	
	for i, file_path in enumerate(transcript_files, 1):
		click.echo(f"\n[{i}/{len(transcript_files)}] Processing: {file_path.name}")
		
		try:
			results = analyzer.analyze_transcript(str(file_path))
			
			# Save individual result
			output_path = file_path.with_suffix('.results.json')
			analyzer.save_results(results, str(output_path))
			
			# Add to batch results with Phase 2B metrics
			batch_record = {
				'file': file_path.name,
				'company': results.company_name,
				'quarter': results.quarter,
				'year': results.year,
				'sentiment_score': results.overall_sentiment.hybrid_sentiment_score,
				'complexity_score': results.overall_complexity.composite_score,
				'transparency_score': results.overall_numerical.numeric_transparency_score,
				'deception_risk_score': results.deception_risk.overall_risk_score if results.deception_risk else None,
				'evasiveness_score': results.evasiveness_scores.overall_evasiveness if results.evasiveness_scores else None,
			}

			# Add Phase 2B metrics if available
			if results.sentence_density_metrics:
				sdm = results.sentence_density_metrics
				batch_record.update({
					'total_sentences': sdm.total_sentences,
					'dense_sentences': sdm.numeric_dense_sentences,
					'proportion_dense': round(sdm.proportion_numeric_dense * 100, 2),
					'mean_density': round(sdm.mean_numeric_density, 2),
				})

			if results.distribution_patterns:
				dp = results.distribution_patterns
				batch_record.update({
					'pattern_type': dp.pattern_type,
					'pattern_confidence': round(dp.pattern_confidence * 100, 1),
					'beginning_density': round(dp.beginning_density, 2),
					'middle_density': round(dp.middle_density, 2),
					'end_density': round(dp.end_density, 2),
					'qa_differential': round(dp.qa_density_differential, 2),
				})

			if results.informativeness_metrics:
				im = results.informativeness_metrics
				batch_record.update({
					'nir': round(im.numeric_inclusion_ratio * 100, 2),
					'informativeness_score': round(im.informativeness_score, 1),
					'forecast_relevance': round(im.forecast_relevance_score, 1),
					'transparency_tier': im.transparency_tier,
				})

			results_list.append(batch_record)
			
		except Exception as e:
			click.echo(f"  ❌ Error: {str(e)}")
			continue
		
	# Save batch summary
	batch_output = dir_path / f"batch_results.{format}"
	
	if format == 'json':
		with open(batch_output, 'w') as f:
			json.dump(results_list, f, indent=2)
	elif format == 'csv':
		import csv
		with open(batch_output, 'w', newline='') as f:
			if results_list:
				writer = csv.DictWriter(f, fieldnames=results_list[0].keys())
				writer.writeheader()
				writer.writerows(results_list)
				
	click.echo(f"\n✓ Batch processing complete!")
	click.echo(f"Summary saved to: {batch_output}")


@cli.command()
@click.argument('input_file', type=click.Path(exists=True), required=False)
@click.option('--output', '-o', help='Output file path for formatted transcript', default=None)
@click.option('--validate-only', is_flag=True, help='Only validate, do not save')
def prepare(input_file, output, validate_only):
	"""
	Interactive transcript preparation wizard for non-technical users

	Guides you through:
	- Metadata entry (company, ticker, quarter, year, date)
	- Speaker identification and correction
	- Section validation
	- Format verification

	Example:
		earnings-analyzer prepare raw_transcript.txt -o formatted.txt
		earnings-analyzer prepare  # Interactive mode from scratch
	"""
	from src.core.transcript_processor import TranscriptProcessor, TranscriptMetadata
	import re
	from datetime import datetime

	click.echo("\n" + "="*70)
	click.echo("📝 TRANSCRIPT PREPARATION WIZARD")
	click.echo("="*70)
	click.echo("\nThis wizard will help you prepare an earnings call transcript")
	click.echo("for analysis. Follow the prompts to provide required information.\n")

	# Step 1: Load or create transcript text
	if input_file:
		click.echo(f"📂 Loading transcript from: {input_file}")
		with open(input_file, 'r', encoding='utf-8') as f:
			transcript_text = f.read()
		click.echo(f"   ✓ Loaded {len(transcript_text)} characters\n")
	else:
		click.echo("📝 No input file provided. You can paste transcript text.\n")
		click.echo("Paste your transcript (press Ctrl+D when done on Unix/Mac, Ctrl+Z then Enter on Windows):")
		import sys
		transcript_text = sys.stdin.read()
		if not transcript_text.strip():
			click.echo("❌ No text provided. Exiting.")
			return

	# Step 2: Collect metadata interactively
	click.echo("\n" + "-"*70)
	click.echo("STEP 1: METADATA ENTRY")
	click.echo("-"*70)

	company_name = click.prompt("Company Name (e.g., 'TechCorp Industries')", type=str)

	# Ticker validation
	while True:
		ticker = click.prompt("Stock Ticker (e.g., 'TECH')", type=str).upper()
		if re.match(r'^[A-Z]{1,5}$', ticker):
			break
		click.echo("   ⚠ Ticker should be 1-5 uppercase letters. Try again.")

	# Quarter selection
	quarter = click.prompt("Quarter", type=click.Choice(['Q1', 'Q2', 'Q3', 'Q4'], case_sensitive=False)).upper()

	# Year validation
	current_year = datetime.now().year
	while True:
		year = click.prompt("Year", type=int, default=current_year)
		if 2000 <= year <= current_year + 1:
			break
		click.echo(f"   ⚠ Year should be between 2000 and {current_year + 1}. Try again.")

	# Date entry (optional but recommended)
	date_str = click.prompt("Call Date (MM/DD/YYYY)", default="", show_default=False)
	if date_str:
		try:
			datetime.strptime(date_str, "%m/%d/%Y")
		except ValueError:
			click.echo("   ⚠ Invalid date format, leaving blank")
			date_str = ""

	click.echo("\n✓ Metadata collected:")
	click.echo(f"   Company: {company_name}")
	click.echo(f"   Ticker: {ticker}")
	click.echo(f"   Quarter: {quarter}")
	click.echo(f"   Year: {year}")
	if date_str:
		click.echo(f"   Date: {date_str}")

	# Step 3: Parse and validate transcript structure
	click.echo("\n" + "-"*70)
	click.echo("STEP 2: TRANSCRIPT VALIDATION")
	click.echo("-"*70)

	processor = TranscriptProcessor()

	# Check if transcript already has metadata header
	has_header = transcript_text.strip().startswith("Company:")

	if has_header:
		click.echo("✓ Transcript appears to have metadata header")
		# Try to parse it
		try:
			temp_path = Path(input_file) if input_file else Path("temp_transcript.txt")
			if not input_file:
				with open(temp_path, 'w', encoding='utf-8') as f:
					f.write(transcript_text)

			parsed = processor.process_file(str(temp_path))

			click.echo(f"✓ Successfully parsed transcript")
			click.echo(f"   - Word count: {parsed.word_count}")
			click.echo(f"   - Sentence count: {parsed.sentence_count}")
			click.echo(f"   - Speakers detected: {len(parsed.speakers)}")
			click.echo(f"   - Sections detected: {len(parsed.sections)}")

			if parsed.speakers:
				click.echo(f"\n   Detected speakers:")
				for speaker in list(parsed.speakers.keys())[:5]:
					click.echo(f"      • {speaker}")
				if len(parsed.speakers) > 5:
					click.echo(f"      ... and {len(parsed.speakers) - 5} more")

			if parsed.sections:
				click.echo(f"\n   Detected sections:")
				for section in parsed.sections.keys():
					click.echo(f"      • {section}")

			# Ask if user wants to update metadata
			if click.confirm("\n   Do you want to update the metadata in the transcript?", default=False):
				has_header = False  # Will rebuild with new metadata
			else:
				# Use existing structure
				if validate_only:
					click.echo("\n✓ Validation complete. Transcript is properly formatted.")
					return

				if output:
					# Just copy to output
					with open(output, 'w', encoding='utf-8') as f:
						f.write(transcript_text)
					click.echo(f"\n✓ Transcript saved to: {output}")
				else:
					click.echo("\n✓ Transcript is ready for analysis.")
					click.echo(f"   Run: python cli.py analyze {input_file}")
				return

		except Exception as e:
			click.echo(f"⚠ Parsing failed: {str(e)}")
			click.echo("   Will help you format the transcript...")
			has_header = False

	# Step 4: Build formatted transcript
	if not has_header:
		click.echo("\n" + "-"*70)
		click.echo("STEP 3: FORMATTING TRANSCRIPT")
		click.echo("-"*70)

		# Build metadata header
		formatted_lines = [
			f"Company: {company_name}",
			f"Ticker: {ticker}",
			f"Quarter: {quarter}",
			f"Year: {year}",
		]

		if date_str:
			formatted_lines.append(f"Date: {date_str}")

		formatted_lines.append("")  # Blank line

		# Check if transcript has section markers
		has_sections = any(marker in transcript_text.upper() for marker in [
			"PREPARED REMARKS", "Q&A", "QUESTIONS AND ANSWERS", "OPERATOR:"
		])

		if has_sections:
			click.echo("✓ Detected section markers in transcript")
			# Just append the content
			formatted_lines.append(transcript_text.strip())
		else:
			click.echo("⚠ No section markers detected")
			click.echo("\nTranscript should have sections like:")
			click.echo("   PREPARED REMARKS")
			click.echo("   QUESTIONS AND ANSWERS")
			click.echo("\nYou can add these manually or the system will treat it as one section.")

			if click.confirm("   Add 'PREPARED REMARKS' header?", default=True):
				formatted_lines.append("PREPARED REMARKS")
				formatted_lines.append("")

			formatted_lines.append(transcript_text.strip())

		final_transcript = "\n".join(formatted_lines)

		# Preview
		click.echo("\n" + "-"*70)
		click.echo("PREVIEW (first 500 characters):")
		click.echo("-"*70)
		click.echo(final_transcript[:500])
		if len(final_transcript) > 500:
			click.echo("...")
		click.echo("-"*70)

		# Validate the formatted transcript
		click.echo("\n⏳ Validating formatted transcript...")

		try:
			# Save to temp file for validation
			temp_path = Path("temp_validation.txt")
			with open(temp_path, 'w', encoding='utf-8') as f:
				f.write(final_transcript)

			parsed = processor.process_file(str(temp_path))

			click.echo("✓ Validation successful!")
			click.echo(f"   - Word count: {parsed.word_count}")
			click.echo(f"   - Sentence count: {parsed.sentence_count}")
			click.echo(f"   - Speakers detected: {len(parsed.speakers)}")

			# Check for potential issues
			warnings = []
			if parsed.word_count < settings.MIN_TRANSCRIPT_LENGTH:
				warnings.append(f"Word count ({parsed.word_count}) is below recommended minimum ({settings.MIN_TRANSCRIPT_LENGTH})")
			if len(parsed.speakers) == 0:
				warnings.append("No speakers detected - make sure to use format 'Name - Title: text'")
			if len(parsed.sections) < 2:
				warnings.append("Only one section detected - consider adding section markers")

			if warnings:
				click.echo("\n⚠ Warnings:")
				for warning in warnings:
					click.echo(f"   • {warning}")

			# Clean up temp file
			temp_path.unlink()

		except Exception as e:
			click.echo(f"❌ Validation failed: {str(e)}")
			click.echo("\nPlease check the format and try again.")
			if temp_path.exists():
				temp_path.unlink()
			return

		# Step 5: Save or display
		if validate_only:
			click.echo("\n✓ Validation complete. Transcript is properly formatted.")
			return

		# Determine output path
		if output:
			output_path = Path(output)
		elif input_file:
			input_path = Path(input_file)
			output_path = input_path.with_stem(input_path.stem + "_formatted")
		else:
			output_path = Path(f"{ticker}_{quarter}{year}_transcript.txt")

		# Save
		with open(output_path, 'w', encoding='utf-8') as f:
			f.write(final_transcript)

		click.echo("\n" + "="*70)
		click.echo("✓ TRANSCRIPT PREPARATION COMPLETE!")
		click.echo("="*70)
		click.echo(f"Formatted transcript saved to: {output_path}")
		click.echo("\nNext steps:")
		click.echo(f"   1. Review the transcript: cat {output_path}")
		click.echo(f"   2. Run analysis: python cli.py analyze {output_path} --summary")
		click.echo("="*70)


@cli.command()
@click.argument('results_file', type=click.Path(exists=True))
@click.option('--output', '-o', help='Output Excel file path', default=None)
@click.option('--include-charts', is_flag=True, default=True, help='Include charts in Excel export')
def export_excel(results_file, output, include_charts):
	"""
	Export analysis results to formatted Excel spreadsheet

	Creates a comprehensive Excel workbook with:
	- Summary dashboard
	- Detailed metrics tables
	- Charts and visualizations
	- Heatmaps and distribution graphics

	Example:
		earnings-analyzer export-excel results.json -o report.xlsx
	"""
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
		from openpyxl.chart import BarChart, LineChart, Reference
		from openpyxl.utils import get_column_letter
	except ImportError:
		click.echo("❌ Error: openpyxl is required for Excel export")
		click.echo("Install it with: pip install openpyxl")
		return

	# Load results
	click.echo(f"\n📊 Loading results from: {results_file}")
	with open(results_file, 'r') as f:
		data = json.load(f)

	# Create workbook
	wb = Workbook()
	wb.remove(wb.active)  # Remove default sheet

	# Define styles
	header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
	header_font = Font(bold=True, color="FFFFFF", size=12)
	subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
	subheader_font = Font(bold=True, size=11)
	title_font = Font(bold=True, size=14)
	border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)

	# Sheet 1: Summary Dashboard
	ws_summary = wb.create_sheet("Summary Dashboard")
	ws_summary.column_dimensions['A'].width = 30
	ws_summary.column_dimensions['B'].width = 20

	row = 1

	# Title
	ws_summary[f'A{row}'] = "EARNINGS CALL ANALYSIS REPORT"
	ws_summary[f'A{row}'].font = title_font
	row += 2

	# Company info
	ws_summary[f'A{row}'] = "Company:"
	ws_summary[f'B{row}'] = data.get('company_name', 'N/A')
	ws_summary[f'A{row}'].font = Font(bold=True)
	row += 1

	ws_summary[f'A{row}'] = "Quarter:"
	ws_summary[f'B{row}'] = f"{data.get('quarter', 'N/A')} {data.get('year', '')}"
	ws_summary[f'A{row}'].font = Font(bold=True)
	row += 2

	# Key Metrics Section
	ws_summary[f'A{row}'] = "KEY METRICS"
	ws_summary[f'A{row}'].font = header_font
	ws_summary[f'A{row}'].fill = header_fill
	ws_summary[f'B{row}'].fill = header_fill
	row += 1

	# Sentiment
	if 'overall_sentiment' in data:
		sent = data['overall_sentiment']
		ws_summary[f'A{row}'] = "Sentiment Score"
		ws_summary[f'B{row}'] = round(sent.get('hybrid_sentiment_score', 0), 2)
		row += 1
		ws_summary[f'A{row}'] = "Sentiment Label"
		ws_summary[f'B{row}'] = sent.get('hybrid_label', 'N/A')
		row += 1

	# Complexity
	if 'overall_complexity' in data:
		comp = data['overall_complexity']
		ws_summary[f'A{row}'] = "Complexity Score"
		ws_summary[f'B{row}'] = round(comp.get('composite_score', 0), 1)
		row += 1
		ws_summary[f'A{row}'] = "Complexity Level"
		ws_summary[f'B{row}'] = comp.get('complexity_level', 'N/A')
		row += 1

	# Numerical Analysis
	if 'overall_numerical' in data:
		num = data['overall_numerical']
		ws_summary[f'A{row}'] = "Numeric Transparency"
		ws_summary[f'B{row}'] = f"{round(num.get('numeric_transparency_score', 0), 2)}%"
		row += 1

	# Phase 2B Metrics
	if 'informativeness_metrics' in data:
		row += 1
		ws_summary[f'A{row}'] = "PHASE 2B: INFORMATIVENESS"
		ws_summary[f'A{row}'].font = header_font
		ws_summary[f'A{row}'].fill = header_fill
		ws_summary[f'B{row}'].fill = header_fill
		row += 1

		im = data['informativeness_metrics']
		ws_summary[f'A{row}'] = "Numeric Inclusion Ratio (NIR)"
		ws_summary[f'B{row}'] = f"{round(im.get('numeric_inclusion_ratio', 0) * 100, 2)}%"
		row += 1

		ws_summary[f'A{row}'] = "Informativeness Score"
		ws_summary[f'B{row}'] = round(im.get('informativeness_score', 0), 1)
		row += 1

		ws_summary[f'A{row}'] = "Forecast Relevance Score"
		ws_summary[f'B{row}'] = round(im.get('forecast_relevance_score', 0), 1)
		row += 1

		ws_summary[f'A{row}'] = "Transparency Tier"
		ws_summary[f'B{row}'] = im.get('transparency_tier', 'N/A')
		row += 1

	# Deception Risk
	if 'deception_risk' in data and data['deception_risk']:
		row += 1
		ws_summary[f'A{row}'] = "DECEPTION RISK"
		ws_summary[f'A{row}'].font = header_font
		ws_summary[f'A{row}'].fill = header_fill
		ws_summary[f'B{row}'].fill = header_fill
		row += 1

		risk = data['deception_risk']
		ws_summary[f'A{row}'] = "Risk Level"
		ws_summary[f'B{row}'] = risk.get('risk_level', 'N/A')

		# Color code risk level
		risk_level = risk.get('risk_level', '').upper()
		if 'HIGH' in risk_level or 'CRITICAL' in risk_level:
			ws_summary[f'B{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
			ws_summary[f'B{row}'].font = Font(color="FFFFFF", bold=True)
		elif 'MEDIUM' in risk_level:
			ws_summary[f'B{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
		elif 'LOW' in risk_level:
			ws_summary[f'B{row}'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
			ws_summary[f'B{row}'].font = Font(color="FFFFFF", bold=True)

		row += 1
		ws_summary[f'A{row}'] = "Risk Score"
		ws_summary[f'B{row}'] = round(risk.get('overall_risk_score', 0), 1)
		row += 1

	# Sheet 2: Sentence Density Details
	if 'sentence_density_metrics' in data:
		ws_density = wb.create_sheet("Sentence Density")
		ws_density.column_dimensions['A'].width = 30
		ws_density.column_dimensions['B'].width = 15

		sdm = data['sentence_density_metrics']

		row = 1
		ws_density[f'A{row}'] = "SENTENCE-LEVEL DENSITY METRICS"
		ws_density[f'A{row}'].font = title_font
		row += 2

		# Add metrics
		metrics_to_show = [
			('Total Sentences', sdm.get('total_sentences')),
			('Dense Sentences (>10%)', sdm.get('numeric_dense_sentences')),
			('Moderate Sentences (5-10%)', sdm.get('numeric_moderate_sentences')),
			('Sparse Sentences (1-5%)', sdm.get('numeric_sparse_sentences')),
			('Narrative Sentences (0%)', sdm.get('narrative_sentences')),
			('', ''),
			('Mean Density (%)', round(sdm.get('mean_numeric_density', 0), 2)),
			('Median Density (%)', round(sdm.get('median_numeric_density', 0), 2)),
			('Max Density (%)', round(sdm.get('max_numeric_density', 0), 2)),
			('Std Dev Density', round(sdm.get('std_numeric_density', 0), 2)),
			('', ''),
			('Proportion Dense', f"{round(sdm.get('proportion_numeric_dense', 0) * 100, 2)}%"),
			('Proportion Narrative', f"{round(sdm.get('proportion_narrative', 0) * 100, 2)}%"),
		]

		for label, value in metrics_to_show:
			ws_density[f'A{row}'] = label
			ws_density[f'B{row}'] = value
			if label and not any(char.isdigit() for char in str(label)):
				ws_density[f'A{row}'].font = Font(bold=True)
			row += 1

		# Add chart if enabled
		if include_charts and sdm.get('numeric_dense_sentences'):
			row += 2
			ws_density[f'A{row}'] = "SENTENCE DISTRIBUTION"
			ws_density[f'A{row}'].font = subheader_font
			row += 1

			# Prepare chart data
			categories = ['Dense\n(>10%)', 'Moderate\n(5-10%)', 'Sparse\n(1-5%)', 'Narrative\n(0%)']
			values = [
				sdm.get('numeric_dense_sentences', 0),
				sdm.get('numeric_moderate_sentences', 0),
				sdm.get('numeric_sparse_sentences', 0),
				sdm.get('narrative_sentences', 0)
			]

			# Write chart data
			for i, (cat, val) in enumerate(zip(categories, values), 1):
				ws_density[f'D{row + i}'] = cat
				ws_density[f'E{row + i}'] = val

			# Create bar chart
			chart = BarChart()
			chart.title = "Sentence Distribution by Numeric Density"
			chart.y_axis.title = "Number of Sentences"

			data_ref = Reference(ws_density, min_col=5, min_row=row+1, max_row=row+4)
			cats_ref = Reference(ws_density, min_col=4, min_row=row+1, max_row=row+4)
			chart.add_data(data_ref, titles_from_data=False)
			chart.set_categories(cats_ref)

			ws_density.add_chart(chart, f'D{row + 7}')

	# Sheet 3: Distribution Patterns
	if 'distribution_patterns' in data:
		ws_dist = wb.create_sheet("Distribution Patterns")
		ws_dist.column_dimensions['A'].width = 30
		ws_dist.column_dimensions['B'].width = 20

		dp = data['distribution_patterns']

		row = 1
		ws_dist[f'A{row}'] = "DISTRIBUTION PATTERNS"
		ws_dist[f'A{row}'].font = title_font
		row += 2

		ws_dist[f'A{row}'] = "Pattern Type"
		ws_dist[f'B{row}'] = dp.get('pattern_type', 'N/A').upper()
		ws_dist[f'B{row}'].font = Font(bold=True, size=12)
		row += 1

		ws_dist[f'A{row}'] = "Pattern Confidence"
		ws_dist[f'B{row}'] = f"{round(dp.get('pattern_confidence', 0) * 100, 1)}%"
		row += 2

		ws_dist[f'A{row}'] = "POSITIONAL DENSITY"
		ws_dist[f'A{row}'].font = subheader_font
		ws_dist[f'A{row}'].fill = subheader_fill
		ws_dist[f'B{row}'].fill = subheader_fill
		row += 1

		positions = [
			('Beginning (First 20%)', dp.get('beginning_density', 0)),
			('Middle (Middle 60%)', dp.get('middle_density', 0)),
			('End (Last 20%)', dp.get('end_density', 0)),
		]

		for label, value in positions:
			ws_dist[f'A{row}'] = label
			ws_dist[f'B{row}'] = f"{round(value, 2)}%"
			row += 1

		row += 1
		ws_dist[f'A{row}'] = "Q&A ANALYSIS"
		ws_dist[f'A{row}'].font = subheader_font
		ws_dist[f'A{row}'].fill = subheader_fill
		ws_dist[f'B{row}'].fill = subheader_fill
		row += 1

		ws_dist[f'A{row}'] = "Q&A Density Differential"
		ws_dist[f'B{row}'] = f"{round(dp.get('qa_density_differential', 0), 2)}%"
		row += 1

		ws_dist[f'A{row}'] = "Question Avg Density"
		ws_dist[f'B{row}'] = f"{round(dp.get('question_avg_density', 0), 2)}%"
		row += 1

		ws_dist[f'A{row}'] = "Answer Avg Density"
		ws_dist[f'B{row}'] = f"{round(dp.get('answer_avg_density', 0), 2)}%"
		row += 1

	# Determine output path
	if output:
		output_path = Path(output)
	else:
		results_path = Path(results_file)
		output_path = results_path.with_suffix('.xlsx')

	# Save workbook
	wb.save(str(output_path))

	click.echo(f"\n✓ Excel report saved to: {output_path}")
	click.echo(f"   Sheets created: {len(wb.sheetnames)}")
	click.echo(f"   Charts included: {include_charts}")


@cli.command()
def config():
	"""Show current configuration"""
	click.echo("\n📋 CONFIGURATION")
	click.echo("="*60)
	click.echo(f"Project Root: {settings.PROJECT_ROOT}")
	click.echo(f"Data Directory: {settings.DATA_DIR}")
	click.echo(f"\nOllama:")
	click.echo(f"  Host: {settings.OLLAMA_HOST}")
	click.echo(f"  Sentiment Model: {settings.SENTIMENT_MODEL}")
	click.echo(f"\nFeature Flags:")
	click.echo(f"  Phase 2A Deception Analysis: {settings.ENABLE_DECEPTION_ANALYSIS}")
	click.echo(f"  Evasiveness Analysis: {settings.ENABLE_EVASIVENESS_ANALYSIS}")
	click.echo(f"  Q&A Analysis: {settings.ENABLE_QA_ANALYSIS}")
	click.echo(f"\nBenchmarks:")
	click.echo(f"  S&P 500 Net Positivity: {settings.SP500_NET_POSITIVITY}")
	click.echo(f"  S&P 500 Numeric Transparency: {settings.SP500_NUMERIC_TRANSPARENCY}%")
	click.echo(f"  S&P 500 Evasiveness Baseline: {settings.SP500_EVASIVENESS_BASELINE}")
	click.echo(f"\nThresholds:")
	click.echo(f"  Deception Risk Warning: {settings.DECEPTION_RISK_WARNING}")
	click.echo(f"  Deception Risk Critical: {settings.DECEPTION_RISK_CRITICAL}")
	click.echo(f"  Question Avoidance Alert: {settings.QUESTION_AVOIDANCE_ALERT}%")
	click.echo("="*60)
	
	
if __name__ == '__main__':
	cli()