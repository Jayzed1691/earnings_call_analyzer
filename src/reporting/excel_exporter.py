"""
Excel Exporter

Generates comprehensive Excel workbooks with:
- Multi-sheet organization
- Formatted tables and headers
- Embedded charts and visualizations
- Historical trend analysis
- Peer comparison data
- Summary statistics

Uses openpyxl for Excel file generation.
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, RadarChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError(
        "Phase 2C dependencies not installed. Run: pip install -r requirements-phase2c.txt"
    )


class ExcelExporter:
    """Exports earnings call analysis to formatted Excel workbooks"""
    
    def __init__(self):
        """Initialize Excel exporter"""
        # Define standard colors
        self.colors = {
            'header': 'FFFFFF',
            'header_bg': '4472C4',
            'subheader_bg': 'B4C7E7',
            'success': '70AD47',
            'warning': 'FFC000',
            'danger': 'FF0000',
            'neutral': '7F7F7F',
            'light_gray': 'F2F2F2'
        }
    
    def export_analysis(
        self,
        company_name: str,
        analysis_data: Dict[str, Any],
        historical_data: Optional[List[Dict[str, Any]]] = None,
        peer_data: Optional[List[Dict[str, Any]]] = None,
        output_path: str = None
    ) -> str:
        """
        Export analysis to Excel workbook
        
        Args:
            company_name: Name of the company
            analysis_data: Current analysis results
            historical_data: Historical analysis data (optional)
            peer_data: Peer comparison data (optional)
            output_path: Path to save Excel file
            
        Returns:
            Path to generated Excel file
        """
        # Set default output path
        if output_path is None:
            output_path = f"/home/claude/analysis_{company_name.replace(' ', '_').lower()}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, company_name, analysis_data)
        self._create_detailed_metrics_sheet(wb, analysis_data)
        self._create_insights_sheet(wb, analysis_data)
        
        if historical_data and len(historical_data) > 0:
            self._create_historical_sheet(wb, company_name, historical_data)
        
        if peer_data and len(peer_data) > 0:
            self._create_peer_comparison_sheet(wb, company_name, analysis_data, peer_data)
        
        # Save workbook
        output_path = str(Path(output_path).with_suffix('.xlsx'))
        wb.save(output_path)
        
        return output_path
    
    def _create_summary_sheet(
        self, wb: Workbook, company_name: str, analysis: Dict[str, Any]
    ):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Header
        ws['A1'] = f"Earnings Call Analysis - {company_name}"
        ws['A1'].font = Font(size=16, bold=True, color=self.colors['header'])
        ws['A1'].fill = PatternFill(start_color=self.colors['header_bg'], 
                                     end_color=self.colors['header_bg'], 
                                     fill_type='solid')
        ws.merge_cells('A1:F1')
        
        # Metadata
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%B %d, %Y %I:%M %p")
        row += 1
        ws[f'A{row}'] = "Quarter:"
        ws[f'B{row}'] = f"{analysis.get('quarter', 'N/A')} {analysis.get('year', 'N/A')}"
        row += 2
        
        # Key Metrics Summary
        ws[f'A{row}'] = "KEY METRICS SUMMARY"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Headers
        headers = ['Metric', 'Score', 'Rating', 'Interpretation', 'Status', 'Notes']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=self.colors['header'])
            cell.fill = PatternFill(start_color=self.colors['header_bg'],
                                    end_color=self.colors['header_bg'],
                                    fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Sentiment
        sentiment = analysis.get('sentiment', {})
        sentiment_score = sentiment.get('hybrid_score', 0)
        ws[f'A{row}'] = "Sentiment"
        ws[f'B{row}'] = f"{sentiment_score:.1f}"
        ws[f'C{row}'] = self._get_rating(sentiment_score, False)
        ws[f'D{row}'] = sentiment.get('label', 'N/A')
        ws[f'E{row}'] = self._get_status_icon(sentiment_score, False)
        ws[f'F{row}'] = "Overall tone and positivity"
        self._apply_score_color(ws[f'B{row}'], sentiment_score, False)
        row += 1
        
        # Deception Risk
        deception = analysis.get('deception_risk', {})
        deception_score = deception.get('overall_risk_score', 0)
        ws[f'A{row}'] = "Deception Risk"
        ws[f'B{row}'] = f"{deception_score:.1f}"
        ws[f'C{row}'] = self._get_rating(deception_score, True)
        ws[f'D{row}'] = deception.get('risk_level', 'N/A')
        ws[f'E{row}'] = self._get_status_icon(deception_score, True)
        ws[f'F{row}'] = "Linguistic deception indicators"
        self._apply_score_color(ws[f'B{row}'], deception_score, True)
        row += 1
        
        # Complexity
        complexity = analysis.get('complexity', {})
        complexity_score = complexity.get('composite_score', 0)
        ws[f'A{row}'] = "Complexity"
        ws[f'B{row}'] = f"{complexity_score:.1f}"
        ws[f'C{row}'] = complexity.get('level', 'N/A')
        ws[f'D{row}'] = f"Grade Level: {complexity.get('flesch_kincaid_grade', 0):.1f}"
        ws[f'E{row}'] = self._get_status_icon(complexity_score, True)
        ws[f'F{row}'] = "Language complexity and readability"
        self._apply_score_color(ws[f'B{row}'], complexity_score, True)
        row += 1
        
        # Transparency
        numerical = analysis.get('numerical_transparency', {})
        transparency_score = numerical.get('transparency_score', 0)
        ws[f'A{row}'] = "Transparency"
        ws[f'B{row}'] = f"{transparency_score:.1f}"
        ws[f'C{row}'] = self._get_rating(transparency_score, False)
        ws[f'D{row}'] = f"Specificity: {numerical.get('specificity_index', 0):.2f}"
        ws[f'E{row}'] = self._get_status_icon(transparency_score, False)
        ws[f'F{row}'] = "Numerical clarity and detail"
        self._apply_score_color(ws[f'B{row}'], transparency_score, False)
        row += 1
        
        # Evasiveness
        evasiveness = analysis.get('evasiveness', {})
        evasiveness_score = evasiveness.get('overall_score', 0)
        ws[f'A{row}'] = "Evasiveness"
        ws[f'B{row}'] = f"{evasiveness_score:.1f}"
        ws[f'C{row}'] = self._get_rating(evasiveness_score, True)
        ws[f'D{row}'] = evasiveness.get('level', 'N/A')
        ws[f'E{row}'] = self._get_status_icon(evasiveness_score, True)
        ws[f'F{row}'] = "Question avoidance patterns"
        self._apply_score_color(ws[f'B{row}'], evasiveness_score, True)
        row += 2
        
        # Document Statistics
        ws[f'A{row}'] = "DOCUMENT STATISTICS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        ws[f'A{row}'] = "Word Count:"
        ws[f'B{row}'] = analysis.get('word_count', 0)
        row += 1
        ws[f'A{row}'] = "Sentence Count:"
        ws[f'B{row}'] = analysis.get('sentence_count', 0)
        row += 1
        ws[f'A{row}'] = "Avg Words/Sentence:"
        ws[f'B{row}'] = f"{analysis.get('word_count', 0) / max(analysis.get('sentence_count', 1), 1):.1f}"
        
        # Format columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 35
        
        # Add chart
        self._add_summary_chart(ws, row + 3)
    
    def _create_detailed_metrics_sheet(self, wb: Workbook, analysis: Dict[str, Any]):
        """Create detailed metrics breakdown sheet"""
        ws = wb.create_sheet("Detailed Metrics")
        
        # Header
        ws['A1'] = "Detailed Metrics Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        row = 3
        
        # Sentiment Details
        row = self._add_metric_section(
            ws, row, "SENTIMENT ANALYSIS",
            analysis.get('sentiment', {}),
            [
                ('hybrid_score', 'Overall Score', '/100'),
                ('label', 'Classification', ''),
                ('lexicon_net_positivity', 'Lexicon Net Positivity', ''),
                ('llm_sentiment_score', 'LLM Assessment', '/100')
            ]
        )
        row += 2
        
        # Complexity Details
        row = self._add_metric_section(
            ws, row, "COMPLEXITY ANALYSIS",
            analysis.get('complexity', {}),
            [
                ('composite_score', 'Composite Score', '/100'),
                ('level', 'Complexity Level', ''),
                ('flesch_reading_ease', 'Flesch Reading Ease', ''),
                ('flesch_kincaid_grade', 'Grade Level', ''),
                ('gunning_fog_index', 'Gunning Fog Index', ''),
                ('smog_index', 'SMOG Index', ''),
                ('coleman_liau_index', 'Coleman-Liau Index', '')
            ]
        )
        row += 2
        
        # Numerical Transparency Details
        row = self._add_metric_section(
            ws, row, "NUMERICAL TRANSPARENCY",
            analysis.get('numerical_transparency', {}),
            [
                ('transparency_score', 'Transparency Score', '/100'),
                ('specificity_index', 'Specificity Index', ''),
                ('forward_looking_density', 'Forward Looking %', '%'),
                ('backward_looking_density', 'Backward Looking %', '%'),
                ('forward_to_backward_ratio', 'Forward/Backward Ratio', ''),
                ('contextualization_quality', 'Context Quality', '/100')
            ]
        )
        row += 2
        
        # Deception Risk Details
        deception = analysis.get('deception_risk', {})
        row = self._add_metric_section(
            ws, row, "DECEPTION RISK ANALYSIS",
            deception,
            [
                ('overall_risk_score', 'Overall Risk Score', '/100'),
                ('risk_level', 'Risk Level', ''),
                ('confidence', 'Confidence', '')
            ]
        )
        
        # Linguistic Markers
        markers = deception.get('linguistic_markers', {})
        row += 1
        ws[f'A{row}'] = "Linguistic Markers:"
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        row += 1
        
        marker_metrics = [
            ('hedging_density', 'Hedging Density', '%'),
            ('qualifier_density', 'Qualifier Density', '%'),
            ('passive_voice_percentage', 'Passive Voice', '%'),
            ('pronoun_distancing_percentage', 'Pronoun Distancing', '%')
        ]
        
        for key, label, unit in marker_metrics:
            ws[f'B{row}'] = f"{label}:"
            value = markers.get(key, 0)
            ws[f'C{row}'] = value
            ws[f'D{row}'] = unit
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 30
    
    def _create_insights_sheet(self, wb: Workbook, analysis: Dict[str, Any]):
        """Create insights and findings sheet"""
        ws = wb.create_sheet("Insights & Findings")
        
        # Header
        ws['A1'] = "Key Insights & Findings"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        insights = analysis.get('insights', {})
        row = 3
        
        # Key Findings
        ws[f'A{row}'] = "KEY FINDINGS"
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.colors['header'])
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['success'],
                                         end_color=self.colors['success'],
                                         fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        findings = insights.get('key_findings', [])
        for finding in findings:
            ws[f'B{row}'] = f"• {finding}"
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        if not findings:
            ws[f'B{row}'] = "No key findings identified"
            row += 1
        
        row += 1
        
        # Red Flags
        ws[f'A{row}'] = "RED FLAGS"
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.colors['header'])
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['danger'],
                                         end_color=self.colors['danger'],
                                         fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        red_flags = insights.get('red_flags', [])
        for flag in red_flags:
            ws[f'B{row}'] = f"⚠ {flag}"
            ws[f'B{row}'].font = Font(color=self.colors['danger'])
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        if not red_flags:
            ws[f'B{row}'] = "No red flags identified"
            row += 1
        
        row += 1
        
        # Strengths
        ws[f'A{row}'] = "STRENGTHS"
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.colors['header'])
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['success'],
                                         end_color=self.colors['success'],
                                         fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        strengths = insights.get('strengths', [])
        for strength in strengths:
            ws[f'B{row}'] = f"✓ {strength}"
            ws[f'B{row}'].font = Font(color=self.colors['success'])
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        if not strengths:
            ws[f'B{row}'] = "No specific strengths highlighted"
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_historical_sheet(
        self, wb: Workbook, company_name: str, historical: List[Dict[str, Any]]
    ):
        """Create historical trend analysis sheet"""
        ws = wb.create_sheet("Historical Trends")
        
        # Header
        ws['A1'] = f"Historical Analysis - {company_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        row = 3
        
        # Column headers
        headers = ['Quarter', 'Year', 'Sentiment', 'Deception Risk', 
                   'Complexity', 'Transparency', 'Evasiveness', 'Word Count']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=self.colors['header'])
            cell.fill = PatternFill(start_color=self.colors['header_bg'],
                                    end_color=self.colors['header_bg'],
                                    fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Sort historical data
        sorted_data = sorted(
            historical,
            key=lambda x: (x.get('year', 0), self._quarter_to_num(x.get('quarter', 'Q1')))
        )
        
        # Data rows
        for data in sorted_data:
            ws[f'A{row}'] = data.get('quarter', 'N/A')
            ws[f'B{row}'] = data.get('year', 'N/A')
            ws[f'C{row}'] = data.get('sentiment', {}).get('hybrid_score', 0)
            ws[f'D{row}'] = data.get('deception_risk', {}).get('overall_risk_score', 0)
            ws[f'E{row}'] = data.get('complexity', {}).get('composite_score', 0)
            ws[f'F{row}'] = data.get('numerical_transparency', {}).get('transparency_score', 0)
            ws[f'G{row}'] = data.get('evasiveness', {}).get('overall_score', 0)
            ws[f'H{row}'] = data.get('word_count', 0)
            
            # Apply formatting
            for col in ['C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].number_format = '0.0'
            
            row += 1
        
        # Add trend chart
        self._add_historical_chart(ws, len(sorted_data) + 4, len(sorted_data))
        
        # Format columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
    
    def _create_peer_comparison_sheet(
        self, wb: Workbook, company_name: str, current: Dict[str, Any], peers: List[Dict[str, Any]]
    ):
        """Create peer comparison sheet"""
        ws = wb.create_sheet("Peer Comparison")
        
        # Header
        ws['A1'] = f"Peer Comparison - {company_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        row = 3
        
        # Column headers
        headers = ['Company', 'Sentiment', 'Deception Risk', 
                   'Complexity', 'Transparency', 'Overall Rating']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=self.colors['header'])
            cell.fill = PatternFill(start_color=self.colors['header_bg'],
                                    end_color=self.colors['header_bg'],
                                    fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Current company (highlighted)
        ws[f'A{row}'] = f"{company_name} (Current)"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = current.get('sentiment', {}).get('hybrid_score', 0)
        ws[f'C{row}'] = current.get('deception_risk', {}).get('overall_risk_score', 0)
        ws[f'D{row}'] = current.get('complexity', {}).get('composite_score', 0)
        ws[f'E{row}'] = current.get('numerical_transparency', {}).get('transparency_score', 0)
        
        # Calculate overall rating
        overall = self._calculate_overall_rating(current)
        ws[f'F{row}'] = overall
        ws[f'F{row}'].font = Font(bold=True)
        
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].number_format = '0.0'
            ws[f'{col}{row}'].fill = PatternFill(start_color=self.colors['light_gray'],
                                                  end_color=self.colors['light_gray'],
                                                  fill_type='solid')
        row += 1
        
        # Peer companies
        for peer in peers:
            ws[f'A{row}'] = peer.get('company_name', 'Unknown')
            ws[f'B{row}'] = peer.get('sentiment', {}).get('hybrid_score', 0)
            ws[f'C{row}'] = peer.get('deception_risk', {}).get('overall_risk_score', 0)
            ws[f'D{row}'] = peer.get('complexity', {}).get('composite_score', 0)
            ws[f'E{row}'] = peer.get('numerical_transparency', {}).get('transparency_score', 0)
            
            overall = self._calculate_overall_rating(peer)
            ws[f'F{row}'] = overall
            
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].number_format = '0.0'
            
            row += 1
        
        # Add comparison chart
        self._add_peer_chart(ws, row + 2, len(peers) + 1)
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    def _add_metric_section(
        self, ws, start_row: int, title: str, data: Dict, metrics: List[tuple]
    ) -> int:
        """Add a metric section to the worksheet"""
        row = start_row
        
        # Section title
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        # Metrics
        for key, label, unit in metrics:
            ws[f'B{row}'] = f"{label}:"
            ws[f'B{row}'].font = Font(bold=True)
            
            value = data.get(key, 'N/A')
            if isinstance(value, (int, float)):
                ws[f'C{row}'] = value
                if unit:
                    ws[f'D{row}'] = unit
            else:
                ws[f'C{row}'] = str(value)
            
            row += 1
        
        return row
    
    def _add_summary_chart(self, ws, start_row: int):
        """Add a summary bar chart to the worksheet"""
        # Chart data is already in the summary section
        # This would add a visual chart (implementation simplified for now)
        pass
    
    def _add_historical_chart(self, ws, start_row: int, data_rows: int):
        """Add historical trend line chart"""
        chart = LineChart()
        chart.title = "Historical Metrics Trend"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Quarter"
        
        # Add data series
        for col, title in enumerate(['Sentiment', 'Deception Risk', 'Complexity'], start=3):
            data = Reference(ws, min_col=col, min_row=3, max_row=3+data_rows)
            chart.add_data(data, titles_from_data=False, from_rows=False)
        
        # Add labels
        labels = Reference(ws, min_col=1, min_row=4, max_row=3+data_rows)
        chart.set_categories(labels)
        
        chart.height = 12
        chart.width = 24
        
        ws.add_chart(chart, f'A{start_row}')
    
    def _add_peer_chart(self, ws, start_row: int, num_companies: int):
        """Add peer comparison bar chart"""
        chart = BarChart()
        chart.title = "Peer Comparison - Key Metrics"
        chart.type = "col"
        chart.grouping = "clustered"
        
        # Add data series
        for col, title in enumerate(['Sentiment', 'Transparency'], start=2):
            data = Reference(ws, min_col=col, min_row=3, max_row=3+num_companies)
            chart.add_data(data, titles_from_data=False)
        
        # Add company labels
        labels = Reference(ws, min_col=1, min_row=4, max_row=3+num_companies)
        chart.set_categories(labels)
        
        chart.height = 12
        chart.width = 24
        
        ws.add_chart(chart, f'A{start_row}')
    
    def _apply_score_color(self, cell, score: float, reverse: bool = False):
        """Apply color based on score"""
        if reverse:
            # Lower is better (e.g., deception risk)
            if score < 40:
                color = self.colors['success']
            elif score < 60:
                color = self.colors['warning']
            else:
                color = self.colors['danger']
        else:
            # Higher is better (e.g., sentiment)
            if score >= 60:
                color = self.colors['success']
            elif score >= 40:
                color = self.colors['warning']
            else:
                color = self.colors['danger']
        
        cell.font = Font(bold=True, color=color)
    
    def _get_rating(self, score: float, reverse: bool = False) -> str:
        """Get rating label based on score"""
        if reverse:
            if score < 40:
                return "Excellent"
            elif score < 60:
                return "Good"
            else:
                return "Concerning"
        else:
            if score >= 60:
                return "Excellent"
            elif score >= 40:
                return "Good"
            else:
                return "Poor"
    
    def _get_status_icon(self, score: float, reverse: bool = False) -> str:
        """Get status icon based on score"""
        if reverse:
            if score < 40:
                return "✓"
            elif score < 60:
                return "○"
            else:
                return "✗"
        else:
            if score >= 60:
                return "✓"
            elif score >= 40:
                return "○"
            else:
                return "✗"
    
    def _calculate_overall_rating(self, data: Dict[str, Any]) -> float:
        """Calculate overall rating from multiple metrics"""
        sentiment = data.get('sentiment', {}).get('hybrid_score', 0)
        transparency = data.get('numerical_transparency', {}).get('transparency_score', 0)
        deception = 100 - data.get('deception_risk', {}).get('overall_risk_score', 0)
        complexity = 100 - data.get('complexity', {}).get('composite_score', 0)
        
        # Weighted average
        overall = (sentiment * 0.3 + transparency * 0.3 + deception * 0.25 + complexity * 0.15)
        return overall
    
    def _quarter_to_num(self, quarter: str) -> int:
        """Convert quarter string to number"""
        quarter_map = {'Q1': 1, 'Q2': 2, 'Q3': 3, 'Q4': 4}
        return quarter_map.get(quarter, 0)


def main():
    """Example usage"""
    
    # Example: Generate a sample Excel report
    exporter = ExcelExporter()
    
    # Mock analysis data
    sample_data = {
        'sentiment': {
            'hybrid_score': 65.5,
            'label': 'Positive',
            'lexicon_net_positivity': 0.15,
            'llm_sentiment_score': 68.2
        },
        'complexity': {
            'composite_score': 55.8,
            'level': 'Moderate',
            'flesch_reading_ease': 45.2,
            'flesch_kincaid_grade': 12.5,
            'gunning_fog_index': 14.2,
            'smog_index': 13.1,
            'coleman_liau_index': 12.8
        },
        'numerical_transparency': {
            'transparency_score': 72.3,
            'specificity_index': 0.78,
            'forward_looking_density': 3.2,
            'backward_looking_density': 5.1,
            'forward_to_backward_ratio': 0.63,
            'contextualization_quality': 75.0
        },
        'deception_risk': {
            'overall_risk_score': 32.5,
            'risk_level': 'Low',
            'confidence': 0.85,
            'linguistic_markers': {
                'hedging_density': 7.5,
                'qualifier_density': 5.2,
                'passive_voice_percentage': 18.2,
                'pronoun_distancing_percentage': 12.3
            }
        },
        'evasiveness': {
            'overall_score': 25.8,
            'level': 'Low'
        },
        'insights': {
            'key_findings': [
                'Strong revenue growth of 15% YoY',
                'Positive forward guidance provided'
            ],
            'red_flags': [],
            'strengths': [
                'Clear and transparent communication',
                'Detailed numerical context'
            ]
        },
        'quarter': 'Q4',
        'year': 2024,
        'word_count': 5234,
        'sentence_count': 287
    }
    
    output = exporter.export_analysis(
        company_name="Example Corp",
        analysis_data=sample_data,
        output_path="/home/claude/sample_report.xlsx"
    )
    
    print(f"Excel report generated: {output}")


if __name__ == "__main__":
    main()
